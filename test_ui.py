import pytest
from commons.excelOpration import ExcelOpration
from commons.oprationElement import WebDriverWrapper
import openpyxl
import os
from pathlib import Path


#从ui_test_cases.xlsx获取测试用例集
def get_case_list():
    # 自动获取当前 py 文件所在的目录
    BASE_DIR = Path(__file__).parent

    # 拼接相对路径
    excel_path = BASE_DIR / "testdata" / "ui_test_cases.xlsx"

    # 1. 打开Excel文件
    wb = openpyxl.load_workbook(excel_path)

    # 2. 选择第 3 个sheet（sheet=2 对应索引2，openpyxl从0开始）
    sheet = wb.worksheets[2]

    case_list = []
    headers = []

    # 3. 读取第一行作为key
    for col in sheet[1]:
        headers.append(col.value)

    # 4. 从第二行开始读取数据，转成字典
    for row in list(sheet.rows)[1:]:
        row_data = {}
        for i, cell in enumerate(row):
            row_data[headers[i]] = cell.value
        case_list.append(row_data)

    # 5. 关闭文件
    wb.close()

    print(f'case_list={case_list}')
    return case_list #返回测试用例集

class TestUiCases():

    #准备工作
    def setup_class(self):

        # 读取ui_test_cases的TestConfig配置
        self.wb=openpyxl.load_workbook(filename='testdata/ui_test_cases.xlsx',read_only=False)

        # 获取sheet_Testconfig
        self.sheet_Testconfig = self.wb["TestConfig"]
        # 获取sheet_TestSteps
        self.sheet_TestSteps = self.wb["TestSteps"]
        # 获取sheet_PageElements
        self.sheet_PageElements = self.wb["PageElements"]
        # 获取单元格对象-driver
        self.driver_path = self.sheet_Testconfig["B1"].value
        print('driver_path----'+str(self.driver_path))
        # 获取单元格对象-url
        self.url = self.sheet_Testconfig["B2"].value
        print('url----' + str(self.url))
        #获取webdriver
        self.driver = WebDriverWrapper(self.driver_path)
        self.driver.get(self.url)

        # 读取ui_test_cases的TestSuite数据
        self.sheet_TestSuite = self.wb["TestSuite"]
        self.excel_Opration = ExcelOpration()
        rows = self.excel_Opration.count_rows(self.sheet_TestSuite)
        # 获取总的用例个数：
        print('\n总的用例个数： ' + str(rows))

        # 获取总的需要自动化执行在用例个数：
        auto_count = self.excel_Opration.auto_count_rows(self.sheet_TestSuite, 'C')
        print('\n总的需要自动化执行的用例个数： ' + str(auto_count))


    # @pytest.mark.parametrize('case_dict',get_case_list())
    @pytest.mark.parametrize('idx, case_dict', enumerate(get_case_list()))
    def test_exe(self,idx,case_dict): #idx是当前执行的第几个用例
        print(f"idx={idx}")

        #遍历读取ui_test_cases的TestSuite的需要自动化执行的测试用例序号;去TestSteps根据测试用例序号找到对应在测试用例步骤去执行
        j = 0  # 计数自动化用例的个数
        # for i in range(1,rows+2):

            # print(f'j={j}')
        # if self.sheet_TestSuite.cell(row=i,column=6).value=='yes':
        if case_dict.get('是否执行') == 'yes':
            j=j+1
            #获取当前的测试用例序号
            self.caseNo=case_dict.get('测试用例序号')
            # print(f'j={j}')
            # 找到需要执行在测试用例编号
            # self.caseNo = self.sheet_TestSuite.cell(row=i,column=1).value
            print(f'当前测试NO.: {self.caseNo}')
            print(f'执行第 {j} 测试用例: {self.caseNo}')

            #去TestSteps找到对应在测试用例步骤
            # 遍历TestSteps表列A中的每个单元格
            step_list=[]
            for cell in self.sheet_TestSteps['A']:
                if cell.value == self.caseNo:
                    step_list.append(cell.row)
            print(f'当前测试用例的行号是{step_list}')

            #获取TestSteps表列B的页面元素
            for k in step_list:
                m=0#为0默认执行成功，1执行不成功
                element=self.sheet_TestSteps.cell(row=k,column=3).value
                # print(f'element={element}')
                print(f'执行测试步骤第{k}行：{element}')
                element_value=self.sheet_TestSteps.cell(row=k,column=4).value
                # print(f'element_value={element_value}')
                verify_value = self.sheet_TestSteps.cell(row=k, column=5).value
                # print(f'verify_value={verify_value}')
                #去pageElement找到对应在元素，获取元素相关信息
                get_element=self.excel_Opration.get_element_info(element,self.sheet_PageElements,'B')
                # print(f'获取元素信息：{get_element}')
                #对获取到的元素操作
                opra=get_element[0]
                # print(f'opra={opra}')
                get_method=get_element[1]
                # print(f'get_method={get_method}')
                opra_value=get_element[2]
                # print(f'opra={opra_value}')
                if opra=='input_text':
                    m=self.driver.input_text(get_method,opra_value,element_value)
                    # print(f'm={m}')
                    if m==0:#执行成功，执行成功写入ui_test_cases_TestSteps F列
                        self.sheet_TestSteps.cell(row=k,column=6).value = '用例步骤执行成功'
                        print(f'执行成功')

                        # self.sheet_TestSteps['F2']='用例步骤执行成功'
                    elif m==1:
                        self.sheet_TestSteps.cell(row=k,column=6).value='用例步骤执行失败'
                        print(f'执行失败，当前案例的后续步骤不执行退出')

                        break #步骤失败退出，不执行后面的步骤

                elif opra=='click_element':
                    m=self.driver.click_element(get_method,opra_value)
                    if m==0:#执行成功，执行成功写入ui_test_cases_TestSteps F列
                        self.sheet_TestSteps.cell(row=k, column=6).value = '用例步骤执行成功'
                        print(f'执行成功')

                    elif m==1:
                        self.sheet_TestSteps.cell(row=k,column=6).value='用例步骤执行失败'
                        print(f'执行失败，当前案例的后续步骤不执行退出')

                        break #步骤失败退出，不执行后面的步骤

                elif opra=='assert':
                    get_text=self.driver.get_element_text(get_method, opra_value)
                    # print(f'get_text={get_text}')
                    assert get_text == verify_value
                    if get_text==verify_value:
                        self.sheet_TestSteps.cell(row=k, column=6).value = '用例步骤执行成功'
                        print(f'测试步骤{k}的检查点期望值：{verify_value} = 实际值：{get_text}')

                    else:
                        self.sheet_TestSteps.cell(row=k, column=6).value = '用例步骤执行失败'
                        print(f'测试步骤的检查点期望值：{verify_value} != 实际值：{get_text}')


            #标注TestSuite的测试用例执行结果
            if m==0: #用例步骤全部执行成功
                self.sheet_TestSuite.cell(row=idx+2,column=4).value='用例执行成功'

            else:
                self.sheet_TestSuite.cell(row=idx+2,column=4).value='用例执行失败'

        self.wb.save(filename="testdata/ui_test_cases.xlsx")

if __name__ == '__main__':
    pytest.main(['-vs', "./test_ui.py", "--alluredir", "./reports/allure-ui-result"])
    os.system(r"allure generate --clean ./reports/allure-ui-result -o ./reports/allure-ui-report")










